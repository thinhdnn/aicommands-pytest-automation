import ast
import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================
# Models
# =========================

class TestStep:
    def __init__(self, action: str, expected: str):
        self.action = action
        self.expected = expected


class TestCase:
    def __init__(
        self,
        tc_id: str,
        feature: str,
        scenario: str,
        test_type: str,
        source: str,
        markers: List[str],
    ):
        self.tc_id = tc_id
        self.feature = feature
        self.scenario = scenario
        self.test_type = test_type
        self.source = source
        self.markers = markers
        self.steps: List[TestStep] = []


# =========================
# AST Parser
# =========================

class TestVisitor(ast.NodeVisitor):
    def __init__(self, file_path: Path, tests_root: Path):
        self.file_path = file_path
        # Base directory to which source paths should be relative (the `tests` folder)
        self.tests_root = tests_root
        self.test_cases: List[TestCase] = []
        self.current_test: TestCase | None = None
        self.pending_step: str | None = None

    def visit_FunctionDef(self, node: ast.FunctionDef):
        markers = []
        for deco in node.decorator_list:
            if isinstance(deco, ast.Attribute):
                markers.append(deco.attr)
            elif isinstance(deco, ast.Call) and isinstance(deco.func, ast.Attribute):
                markers.append(deco.func.attr)

        test_type = "UI" if "ui" in markers else "API"
        feature = markers[1] if len(markers) > 1 else "generated"

        tc_id = f"{feature.upper()}_{node.name.upper()}"
        # Compute source path relative to the tests root for cleaner reporting
        try:
            relative_source = self.file_path.relative_to(self.tests_root)
        except ValueError:
            # Fallback: if file is unexpectedly outside tests_root, use basename
            relative_source = self.file_path.name

        self.current_test = TestCase(
            tc_id=tc_id,
            feature=feature,
            scenario=node.name,
            test_type=test_type,
            source=str(relative_source),
            markers=markers,
        )

        self.generic_visit(node)

        if self.current_test.steps:
            self.test_cases.append(self.current_test)

        self.current_test = None

    def visit_Call(self, node: ast.Call):
        if not self.current_test:
            return

        # test.step("...")
        if (
            isinstance(node.func, ast.Attribute)
            and node.func.attr == "step"
            and node.args
            and isinstance(node.args[0], ast.Constant)
        ):
            self.pending_step = node.args[0].value

        # .expect("...")
        if (
            isinstance(node.func, ast.Attribute)
            and node.func.attr == "expect"
            and node.args
            and isinstance(node.args[0], ast.Constant)
            and self.pending_step
        ):
            self.current_test.steps.append(
                TestStep(
                    action=self.pending_step,
                    expected=node.args[0].value,
                )
            )
            self.pending_step = None

        self.generic_visit(node)


# =========================
# Collect tests
# =========================

def collect_test_cases(test_root: Path) -> List[TestCase]:
    cases: List[TestCase] = []

    for py_file in test_root.rglob("test_*.py"):
        tree = ast.parse(py_file.read_text(encoding="utf-8"))
        # Ensure sources are reported relative to the shared `tests` directory
        visitor = TestVisitor(py_file, tests_root=test_root.parent)
        visitor.visit(tree)
        cases.extend(visitor.test_cases)

    return cases


# =========================
# Markdown output
# =========================

def write_markdown(cases: List[TestCase], output: Path):
    lines = [
        "# Test Cases",
        "",
        "_Auto-generated from test scripts. Do not edit manually._",
        "",
    ]

    for tc in cases:
        lines.extend([
            f"## {tc.tc_id}",
            f"- **Feature**: {tc.feature}",
            f"- **Scenario**: {tc.scenario}",
            f"- **Type**: {tc.test_type}",
            f"- **Source**: `{tc.source}`",
            "",
            "| # | Step | Expected Result |",
            "|---|------|-----------------|",
        ])

        for i, step in enumerate(tc.steps, 1):
            lines.append(f"| {i} | {step.action} | {step.expected} |")

        lines.append("")

    output.write_text("\n".join(lines), encoding="utf-8")


# =========================
# Excel output
# =========================

def write_excel(cases: List[TestCase], output: Path):
    def read_existing_rows_by_tc_id(path: Path) -> Dict[str, Dict[str, object]]:
        """Read existing TestCases sheet and return a mapping TC_ID -> row dict.

        This is used to preserve manual fields (Priority/Status/Owner/etc.) across regenerations
        while still updating auto-generated fields (Steps/Expected/Markers/Source).
        """
        if not path.exists():
            return {}

        try:
            existing_wb = load_workbook(path)
        except Exception:
            return {}

        if "TestCases" not in existing_wb.sheetnames:
            return {}

        existing_ws = existing_wb["TestCases"]
        if existing_ws.max_row < 2:
            return {}

        existing_headers: List[str] = []
        for cell in existing_ws[1]:
            existing_headers.append(str(cell.value).strip() if cell.value is not None else "")

        header_to_col: Dict[str, int] = {
            h: i + 1 for i, h in enumerate(existing_headers) if h
        }
        tc_col = header_to_col.get("TC_ID")
        if not tc_col:
            return {}

        rows: Dict[str, Dict[str, object]] = {}
        for r in range(2, existing_ws.max_row + 1):
            raw_tc_id = existing_ws.cell(row=r, column=tc_col).value
            if raw_tc_id is None:
                continue
            tc_id = str(raw_tc_id).strip()
            if not tc_id:
                continue

            row_dict: Dict[str, object] = {}
            for h, c in header_to_col.items():
                row_dict[h] = existing_ws.cell(row=r, column=c).value
            rows[tc_id] = row_dict

        return rows

    existing_rows = read_existing_rows_by_tc_id(output)

    wb = Workbook()

    # Summary sheet (manager-friendly overview)
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Detailed test case sheet
    ws = wb.create_sheet("TestCases")

    headers = [
        "TC_ID",
        "Title",
        "Feature/Module",
        "Scenario (function)",
        "Type",
        "Priority",
        "Status",
        "Automation",
        "Owner",
        "Preconditions",
        "Steps",
        "Expected Results",
        "Postconditions",
        "Automation Source",
        "Markers/Tags",
        "Last Generated (UTC)",
    ]
    ws.append(headers)

    timestamp = datetime.utcnow().replace(microsecond=0).isoformat()

    manual_columns = {
        "Priority",
        "Status",
        "Automation",
        "Owner",
        "Preconditions",
        "Postconditions",
    }

    generated_tc_ids: set[str] = set()

    for tc in cases:
        steps = "\n".join(f"{i+1}. {s.action}" for i, s in enumerate(tc.steps))
        expected = "\n".join(f"{i+1}. {s.expected}" for i, s in enumerate(tc.steps))
        title = tc.scenario.replace("_", " ").strip().title()

        # Display as `tests/...` while keeping it relative (not absolute)
        automation_source_display = f"tests/{tc.source}"

        generated_tc_ids.add(tc.tc_id)

        base_row = {
            "TC_ID": tc.tc_id,
            "Title": title,
            "Feature/Module": tc.feature,
            "Scenario (function)": tc.scenario,
            "Type": tc.test_type,
            "Priority": "",
            "Status": "Ready",
            "Automation": "Automated",
            "Owner": "",
            "Preconditions": "",
            "Steps": steps,
            "Expected Results": expected,
            "Postconditions": "",
            "Automation Source": automation_source_display,
            "Markers/Tags": ", ".join(tc.markers),
            "Last Generated (UTC)": timestamp,
        }

        # Preserve manually edited fields if the TC already exists in the Excel
        existing = existing_rows.get(tc.tc_id)
        if existing:
            for col in manual_columns:
                val = existing.get(col)
                if val not in (None, ""):
                    base_row[col] = val

        ws.append([
            base_row["TC_ID"],
            base_row["Title"],
            base_row["Feature/Module"],
            base_row["Scenario (function)"],
            base_row["Type"],
            base_row["Priority"],
            base_row["Status"],
            base_row["Automation"],
            base_row["Owner"],
            base_row["Preconditions"],
            base_row["Steps"],
            base_row["Expected Results"],
            base_row["Postconditions"],
            base_row["Automation Source"],
            base_row["Markers/Tags"],
            base_row["Last Generated (UTC)"],
        ])

    # Carry forward rows that exist in the Excel but are not present in the current code scan
    # (e.g., manual test cases or temporarily removed automation).
    for tc_id, row in existing_rows.items():
        if tc_id in generated_tc_ids:
            continue

        ws.append([
            row.get("TC_ID", tc_id),
            row.get("Title", ""),
            row.get("Feature/Module", ""),
            row.get("Scenario (function)", ""),
            row.get("Type", ""),
            row.get("Priority", ""),
            row.get("Status", ""),
            row.get("Automation", ""),
            row.get("Owner", ""),
            row.get("Preconditions", ""),
            row.get("Steps", ""),
            row.get("Expected Results", ""),
            row.get("Postconditions", ""),
            row.get("Automation Source", ""),
            row.get("Markers/Tags", ""),
            row.get("Last Generated (UTC)", ""),
        ])

    # -------------------------
    # Formatting / UX
    # -------------------------
    # Modern professional color scheme
    header_fill = PatternFill("solid", fgColor="2E75B6")  # Professional blue
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Better borders - medium for definition
    medium_side = Side(style="medium", color="4472C4")
    thin_side = Side(style="thin", color="D0D0D0")
    header_border = Border(
        left=medium_side, right=medium_side, top=medium_side, bottom=medium_side
    )
    cell_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # Alignment styles
    wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    plain_top = Alignment(vertical="top", horizontal="left")
    center_align = Alignment(vertical="center", horizontal="center")

    # Header row styling - taller and more prominent
    ws.row_dimensions[1].height = 30
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # Optimized column widths for better readability
    column_widths = {
        "A": 32,  # TC_ID
        "B": 40,  # Title
        "C": 20,  # Feature
        "D": 45,  # Scenario
        "E": 12,  # Type
        "F": 12,  # Priority
        "G": 14,  # Status
        "H": 14,  # Automation
        "I": 20,  # Owner
        "J": 30,  # Preconditions
        "K": 50,  # Steps
        "L": 50,  # Expected
        "M": 30,  # Postconditions
        "N": 50,  # Automation Source
        "O": 32,  # Markers
        "P": 24,  # Timestamp
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header and enable filtering
    ws.freeze_panes = "A2"
    last_row = ws.max_row
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Data validation dropdowns (edit-friendly for humans)
    if last_row >= 2:
        dv_type = DataValidation(type="list", formula1='"API,UI"', allow_blank=True)
        dv_priority = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
        dv_status = DataValidation(type="list", formula1='"Draft,Ready,Blocked,Deprecated"', allow_blank=True)
        dv_automation = DataValidation(type="list", formula1='"Automated,Manual,Partial"', allow_blank=True)

        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_priority)
        ws.add_data_validation(dv_status)
        ws.add_data_validation(dv_automation)

        dv_type.add(f"E2:E{last_row}")
        dv_priority.add(f"F2:F{last_row}")
        dv_status.add(f"G2:G{last_row}")
        dv_automation.add(f"H2:H{last_row}")

    # Row styling & borders with better visual appeal
    even_row_fill = PatternFill("solid", fgColor="F2F2F2")  # Light gray for even rows
    odd_row_fill = PatternFill("solid", fgColor="FFFFFF")  # White for odd rows
    
    # Text font for data rows
    data_font = Font(size=10, color="000000")
    
    for r in range(2, last_row + 1):
        # Auto-adjust row height based on content (minimum 60, max 120)
        ws.row_dimensions[r].height = 60
        
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = cell_border
            cell.font = data_font

            # Alternating row fill for better readability
            if r % 2 == 0:
                cell.fill = even_row_fill
            else:
                cell.fill = odd_row_fill

            # Column-specific alignment
            if c in {5, 6, 7, 8}:  # Type, Priority, Status, Automation - center aligned
                cell.alignment = wrap_center
            elif c in {10, 11, 12, 13, 14, 15}:  # Preconditions, Steps, Expected, Postconditions, Source, Tags - wrap
                cell.alignment = wrap_top
            else:
                cell.alignment = plain_top

        # Hyperlink Automation Source to the file (relative from docs/)
        source_cell = ws.cell(row=r, column=14)
        source_value = source_cell.value
        if isinstance(source_value, str) and source_value.startswith("tests/"):
            source_cell.hyperlink = f"../{source_value}"
            source_cell.font = Font(size=10, color="0563C1", underline="single")

    # Excel table styling - use a more modern style
    if last_row >= 2:
        table_ref = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName="TestCasesTable", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    # -------------------------
    # Summary sheet content - Enhanced formatting
    # -------------------------
    total = len(cases)
    by_type: Dict[str, int] = {}
    by_feature: Dict[str, int] = {}
    for tc in cases:
        by_type[tc.test_type] = by_type.get(tc.test_type, 0) + 1
        by_feature[tc.feature] = by_feature.get(tc.feature, 0) + 1

    # Title with better styling
    ws_summary["A1"].value = "Test Case Summary"
    ws_summary["A1"].font = Font(bold=True, size=18, color="2E75B6")
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    # Info section with better formatting
    info_fill = PatternFill("solid", fgColor="E7F3FF")
    info_border = Border(
        left=Side(style="thin", color="4472C4"),
        right=Side(style="thin", color="4472C4"),
        top=Side(style="thin", color="4472C4"),
        bottom=Side(style="thin", color="4472C4"),
    )
    
    ws_summary["A3"].value = "Last Generated (UTC)"
    ws_summary["A3"].font = Font(bold=True, size=11)
    ws_summary["A3"].fill = info_fill
    ws_summary["A3"].border = info_border
    ws_summary["B3"].value = timestamp
    ws_summary["B3"].fill = info_fill
    ws_summary["B3"].border = info_border
    ws_summary["B3"].alignment = center_align
    
    ws_summary["A4"].value = "Total Test Cases"
    ws_summary["A4"].font = Font(bold=True, size=11)
    ws_summary["A4"].fill = info_fill
    ws_summary["A4"].border = info_border
    ws_summary["B4"].value = total
    ws_summary["B4"].font = Font(bold=True, size=12, color="2E75B6")
    ws_summary["B4"].fill = info_fill
    ws_summary["B4"].border = info_border
    ws_summary["B4"].alignment = center_align

    # Section headers with better styling
    section_fill = PatternFill("solid", fgColor="2E75B6")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    
    ws_summary["A6"].value = "By Type"
    ws_summary["A6"].font = section_font
    ws_summary["A6"].fill = section_fill
    ws_summary["A6"].alignment = center_align
    ws_summary.merge_cells("A6:B6")
    ws_summary.row_dimensions[6].height = 25
    
    # Table headers
    header_fill_summary = PatternFill("solid", fgColor="5B9BD5")
    header_font_summary = Font(bold=True, size=11, color="FFFFFF")
    header_border_summary = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )
    
    ws_summary["A7"].value = "Type"
    ws_summary["A7"].font = header_font_summary
    ws_summary["A7"].fill = header_fill_summary
    ws_summary["A7"].border = header_border_summary
    ws_summary["A7"].alignment = center_align
    ws_summary["B7"].value = "Count"
    ws_summary["B7"].font = header_font_summary
    ws_summary["B7"].fill = header_fill_summary
    ws_summary["B7"].border = header_border_summary
    ws_summary["B7"].alignment = center_align
    
    row = 8
    for t, cnt in sorted(by_type.items(), key=lambda x: x[0]):
        ws_summary[f"A{row}"].value = t
        ws_summary[f"B{row}"].value = cnt
        ws_summary[f"A{row}"].border = cell_border
        ws_summary[f"B{row}"].border = cell_border
        ws_summary[f"B{row}"].alignment = center_align
        if row % 2 == 0:
            ws_summary[f"A{row}"].fill = even_row_fill
            ws_summary[f"B{row}"].fill = even_row_fill
        else:
            ws_summary[f"A{row}"].fill = odd_row_fill
            ws_summary[f"B{row}"].fill = odd_row_fill
        row += 1

    row += 1
    ws_summary[f"A{row}"].value = "By Feature/Module"
    ws_summary[f"A{row}"].font = section_font
    ws_summary[f"A{row}"].fill = section_fill
    ws_summary[f"A{row}"].alignment = center_align
    ws_summary.merge_cells(f"A{row}:B{row}")
    ws_summary.row_dimensions[row].height = 25
    
    row += 1
    ws_summary[f"A{row}"].value = "Feature"
    ws_summary[f"A{row}"].font = header_font_summary
    ws_summary[f"A{row}"].fill = header_fill_summary
    ws_summary[f"A{row}"].border = header_border_summary
    ws_summary[f"A{row}"].alignment = center_align
    ws_summary[f"B{row}"].value = "Count"
    ws_summary[f"B{row}"].font = header_font_summary
    ws_summary[f"B{row}"].fill = header_fill_summary
    ws_summary[f"B{row}"].border = header_border_summary
    ws_summary[f"B{row}"].alignment = center_align
    
    row += 1
    for f, cnt in sorted(by_feature.items(), key=lambda x: (-x[1], x[0])):
        ws_summary[f"A{row}"].value = f
        ws_summary[f"B{row}"].value = cnt
        ws_summary[f"A{row}"].border = cell_border
        ws_summary[f"B{row}"].border = cell_border
        ws_summary[f"B{row}"].alignment = center_align
        if row % 2 == 0:
            ws_summary[f"A{row}"].fill = even_row_fill
            ws_summary[f"B{row}"].fill = even_row_fill
        else:
            ws_summary[f"A{row}"].fill = odd_row_fill
            ws_summary[f"B{row}"].fill = odd_row_fill
        row += 1

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    wb.save(output)


# =========================
# Main
# =========================

def main():
    root = Path.cwd()
    test_dirs = [root / "tests" / "ui", root / "tests" / "api"]

    cases: List[TestCase] = []
    for d in test_dirs:
        if d.exists():
            cases.extend(collect_test_cases(d))

    docs = root / "docs"
    docs.mkdir(exist_ok=True)

    write_markdown(cases, docs / "test_cases.md")
    write_excel(cases, docs / "test_cases.xlsx")

    print(f"Generated {len(cases)} test cases.")


if __name__ == "__main__":
    main()
